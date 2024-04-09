import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import sys
import os
from openpyxl import load_workbook

current_dir = os.path.dirname(__file__)
path = os.path.join(current_dir, '')

baseurl = 'https://onninen.pl'
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36'
}


def link_extraction(link_end):
    product_links = []

    for end in link_end:
        r = requests.get(f'https://onninen.pl/produkty/{end}', headers=headers)
        soup = BeautifulSoup(r.content, 'lxml')
        pagination = soup.find('div', class_='sc-1ia3zb9-0 daMDUy')
        print('Sciąganie z ' + end)

        #First site is diffrent than others
        product_list = soup.find_all('div', class_='vrexg1-5 gLcbnB')
        for item in product_list:
            for link in item.find_all('a', href=True):
                product_links.append(baseurl + link['href'])

        #find how many subsites  there are to iterate through them
        if pagination:
            
            spans = pagination.find_all('span', class_='content')
            
            numbers = [int(span.text) for span in spans if span.text.strip()]
            
            pg_num = max(numbers)

            for x in range(1, pg_num - 1):
                r = requests.get(f'https://onninen.pl/produkty/{end}/strona:{x}', headers=headers)
                soup = BeautifulSoup(r.content, 'lxml')
                product_list = soup.find_all('div', class_='vrexg1-5 gLcbnB')
                print('Pracuje ' + str(x / pg_num))
                for item in product_list:
                    for link in item.find_all('a', href=True):
                        product_links.append(baseurl + link['href'])

    # Saving all links
    with open('product_links.txt', 'w') as f:
        for link in product_links:
            f.write("%s\n" % link)

    print(len(product_links))
    return product_links


def extract_product_data(url):
    r = requests.get(url, headers=headers)
    soup = BeautifulSoup(r.content, 'lxml')

    # Extracting data
    name = soup.find('h1').text.strip()

    div = soup.find_all('span', class_='sc-2vamk6-3 bNiquN list-grid-value')
    trade_index = div[0].text.strip()
    ean = div[2].text.strip()
    

    for paragraph in div[1]:
        manufacturer_index = paragraph.text.strip().split(maxsplit=1)[0]
        break

    try:
        price_classes = ['sc-33rfvt-6 dqWswr price', 'sc-33rfvt-6 jrAnwl price', 'sc-33rfvt-6 knzgHL price']
        jm = None

        for price_class in price_classes:
            price_div = soup.find('div', class_=price_class)
            if price_div is not None:
                jm = price_div.text.split('/')[1].strip()
                break
    except:
        print('Error with product: ' + trade_index)
        sys.exit()

   
    print('Saving product:' + name)

    manufacturer_index = 'Oni' + manufacturer_index
    product = {
        'NAME': name,
        'Original_Name': name,
        'category': 'Ogólna',
        'JM': jm,
        'type': 'Towar',
        'trade_index': trade_index,
        'manufacturer_index': manufacturer_index,
        'EAN': ean,

    }

    return product

# Function to remove temp sheets from Excel file
def remove_sheet_from_excel(filename, sheet_name):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
            wb.save(filename)
            print(f"Sheet '{sheet_name}' removed from '{filename}'.")
        else:
            print(f"Sheet '{sheet_name}' not found in '{filename}'.")
    else:
        print(f"File '{filename}' not found.")


def extract_product_data_from_file():
    product_links = []
    #extracted links path
    with open(f'{path}product_links.txt', 'r') as f:
        product_links = f.readlines()
    product_links = [link.strip() for link in product_links]

    products_data = []
    i = 0
    for link in product_links:
        product_data = extract_product_data(link)
        products_data.append(product_data)
        i += 1
        if i % 10 == 0: #in case of too many requests to the server
            print('Saving...')
            df = pd.DataFrame(products_data)
            if i == 10:
                df.to_excel('productsOninem.xlsx', index=False)
            else:
                with pd.ExcelWriter('productsOninem.xlsx', mode='a', engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, header=False, sheet_name=f'Sheet_{i/10}')
                    remove_sheet_from_excel(f'{path}productsOninem.xlsx', f'Sheet_{(i/10)-1}')
            print('Waiting 3 sec...')
            time.sleep(3) #necessary to avoid timeout

    df = pd.DataFrame(products_data)
    if i > 10:
        with pd.ExcelWriter('productsOninemALL.xlsx', mode='a', engine='openpyxl') as writer:
            df.to_excel(writer, index=False, header=False, sheet_name=f'Sheet_{i/10}')
    else:
        df.to_excel('productsOninemALL.xlsx', index=False, sheet_name='Sheet_1')

#Categories from website that client want to scrap
link_end = ['Technika-instalacyjna', 'Ogrzewanie', 'Kable-i-przewody', 'Sieci-wodno-kanalizacyjne-i-gazowe']
# Extracting product links
product_links = link_extraction(link_end)

# Extracting product data
extract_product_data_from_file()
